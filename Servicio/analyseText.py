import openpyxl
from fuzzywuzzy import fuzz
import re 

def read_dict_tipicacion():
    dictionaryTipification = openpyxl.load_workbook('Tipificacion Dynamics.xlsx')
    hoja1 = dictionaryTipification.get_sheet_by_name('Hoja1')

    motivoConsulta=[]
    negocio=[]
    tipologia3=[]
    tipologia4=[]
    CPC_Escala=[]
    for indexCell in range(1,hoja1.max_row):
        motivoConsulta.append(hoja1.cell(row=indexCell,column = 1).value)
        #print(hoja1.cell(row=indexCell,column = 4).value)
        negocio.append(hoja1.cell(row=indexCell,column = 2).value)
        tipologia3.append(hoja1.cell(row=indexCell,column = 3).value)
        tipologia4.append(hoja1.cell(row=indexCell,column = 4).value)
        CPC_Escala.append(hoja1.cell(row=indexCell,column = 6).value)

    return motivoConsulta,negocio,tipologia3,tipologia4,CPC_Escala

def clean_message(message):
    #messageList=message.split("\n")
    messageList=message.split("   ")
    messageFilter=[]
    for line in messageList:
        lineNoSpaces = re.sub("\W","",line)
        if lineNoSpaces != "":
            messageFilter.append(line)
            #print(line)

    messageBody=""
    flag = 0
    for index_line,line in enumerate(messageFilter):
        #print(str(index_line)+"_________________________"+line)
        if "body" in line:
            if ">" in line:
                flag=1
                messageBody=line
        elif flag == 1:
            if "completado por" not in line.lower():
                messageBody = messageBody +" "+line
                
            else:
                flag =0

    print("-----------------------------------------------------------------------")
    print(messageBody)
    print("-----------------------------------------------------------------")
    return messageBody

def replace_accent(text):
    # Quita las tildes del texto que ingresa
    text = text.replace("á","a")
    text = text.replace("é","e")
    text = text.replace("í","i")
    text = text.replace("ó","o")
    text = text.replace("ú","u")
    text = text.replace(","," ")
    # text = text.replace("."," ")
    # text = text.replace(":"," ")
    text = text.replace(";"," ")
    text = text.replace("body>","")


    return text

def extract_email(messageBody):
    # Recive el cuerpo del mensaje
    # Filtra informacion innecesaria
    # Retorna email
    email="" # Se utiliza cuando no se encuentra el correo en la casilla de email de la pagina
    indexEmail=0
    indexCom=0
    messageBody=messageBody.lower()
    if "email:" in messageBody:
        if ".com " in messageBody:
            indexEmail=messageBody.find("email:")
            indexCom=messageBody.find(".com ")
            email=messageBody[indexEmail+len("email:"):indexCom+len(".com ")]

    elif "correo electronico:" in messageBody:
        if ".com" in messageBody:
            indexEmail=messageBody.find("correo electronico:")
            indexCom=messageBody.find(".com-")
            email=messageBody[indexEmail+len("correo electronico:"):indexCom+len(".com")]
            
    email=re.sub(" ","",email)

    return email


def compare_text(message):
    motivo,negocio,tipologia3,tipologia4,CPC_Escala = read_dict_tipicacion()
    messageBody = clean_message(message)
    messageBody = replace_accent(messageBody)

    # Se tiene en las variables motivo, negocio,tipologia3,tipologia4 las columnas del diccionario dado por Andes
    # Se compara todo el mensaje con el texto de cada celda de la Tipologia4 para lo cual se utiliza fuzzywuzzy
    
    percentajeSimilarity = [] # almacena el porcentaje de coincidencia del texto de cada celda de la tipologia 4 con el mensaje
    for indexT4,statement in enumerate(tipologia4):
        statement = replace_accent(statement)
        #print(statement)
        similarity = fuzz.token_set_ratio(statement,messageBody)
        # print(indexT4,similarity,statement, messageBody)
        #print(indexT4,statement +"________________________"+str(similarity))
        percentajeSimilarity.append(similarity)
        #print(words)
    
    # Se extrae el indice de la tipologia con mayor similitud
    indexMaxSimilarity = percentajeSimilarity.index(max(percentajeSimilarity)) 
    #print(indexMaxSimilarity)

    # Se agrega funcionalidad para extraer correo de los datos de la pagina.
    email=extract_email(messageBody)

    return (motivo[indexMaxSimilarity],negocio[indexMaxSimilarity],tipologia3[indexMaxSimilarity],tipologia4[indexMaxSimilarity],CPC_Escala[indexMaxSimilarity],email)

def tipificacion_message():

    return True

if __name__ == "__main__":
    message = '''  Cargando...              ---              Caso : Caso de servcioHaga clic aquí para seleccionar un formulario diferente para el registro   CHEVIGNON - Pregunta o solicitud.  FormSections_NavigationFlyOut_Button  Registro anterior (Ctrl+<)  Registro siguiente (Ctrl+>)        Required               Select a value.  Control bloqueadoError   Razón para el estado    Seleccione el estado del caso.     Razón para el estado Abierto   Control bloqueadoError    Prioridad    Listado de prioridades     Prioridad    Control bloqueadoError    Número de caso    Muestra el número de caso para tenerlo como referencia del cliente y capacidades de búsqueda. Esto no se puede modificar.     Número de caso Muestra el número de caso para tenerlo como referencia del cliente y capacidades de búsqueda. Esto no se puede modificar.  CAS-255363-M3W6H9    Control bloqueado Error              Cerrar             Contraer esta ficha   Resumen         DETALLES DEL CASO    Título de caso   Recommended Escriba un tema o un nombre descriptivo, como la solicitud, el problema o el nombre de la compañía, para identificar el caso en las vistas de Microsoft Dynamics 365.     Título de caso CHEVIGNON - Pregunta o solicitud.   Control bloqueadoError    Creado Zendex    Creación de cliente de manera automática     Creado Zendex Creación de cliente de manera automática  No    Control bloqueado Error    Origen    Seleccione cómo se originó el contacto sobre el caso, por ejemplo, por correo electrónico, teléfono o web, para su uso en generación de informes y análisis.     Origen Correo   Control bloqueadoError    Marca   Required Referencia al listado de marcas     Marca Chevignon         Select a value.  Control bloqueadoError      Motivo   Required Referencia al listado de tipología 1     Motivo           Select a value.  Control bloqueadoError      Negocio   Required Referencia al listado de tipologías 2     Negocio           Select a value.  Control bloqueadoError      Tipología 3   Required Referencia al listado de tipologías 3     Tipología 3           Select a value.  Control bloqueadoError      Tipología 4   Required Referencia al listado de tipologías 4     Tipología 4           Select a value.  Control bloqueadoError      Tienda    Referencia al listado de tiendas     Tienda           Select a value.  Control bloqueadoError      Estado inicial   Recommended     Estado inicial    Control bloqueadoError    Gestión    Listado de gestiones     Gestión    Control bloqueadoError               Gestión Bodega    Guía Manual TCC    Guía con la que TCC recoge cuando el cliente no imprime la guia.     Guía Manual TCC    Control bloqueadoError   Guia de Cambio    Número de guia con la cual bodega despacha el pedido nuevamente al cliente     Guia de Cambio    Control bloqueadoError    Fecha de recepción CEDI    Fecha con la que los paquetes de online se reciben en CEDI Escriba una fecha con el formato d/MM/yyyy     Fecha de recepción CEDI         Seleccionar fecha  Control bloqueadoError   Nota crédito (DV)    Nota credito con la cual se realizo la devolución     Nota crédito (DV)    Control bloqueadoError    Fecha de Recepción en Bodega    Fecha en la cual se recepcionan las prendas en bodega. Escriba una fecha con el formato d/MM/yyyy     Fecha de Recepción en Bodega         Seleccionar fecha  Control bloqueadoError   Valor nota crédito    Valor de la nota crédito de la devolución     Valor nota crédito    Control bloqueadoError    Fecha Solución Bodega    Fecha en la cual termina la gestión realizada por bodega. Escriba una fecha con el formato d/MM/yyyy     Fecha Solución Bodega         Seleccionar fecha  Control bloqueadoError   Solución Bodega    Motivo de solición bodega     Solución Bodega    Control bloqueadoError                 Cliente    Nombre del cliente    Nombre del cliente     Nombre del cliente    Control bloqueadoError    Celular    Celular del cliente     Celular    Control bloqueadoError    Correo electrónico    La dirección de correo electrónico principal para la entidad. Required    Correo electrónico formulariosweb@gco.com.co    Control bloqueadoError    Tipo de documento    Conjunto de tipo de documentos     Tipo de documento    Control bloqueadoError    No de identificación    No de identificación del cliente     No de identificación    Control bloqueadoError    Ciudad    Ciudad del cliente     Ciudad           Select a value.  Control bloqueadoError      Número de pedido    Número de pedido que origina el caso     Número de pedido    Control bloqueadoError    Guía de despacho    Número de guia con el que se despacho el pedido o se va despachar el pedido.     Guía de despacho    Control bloqueadoError    Referencia    Campo para anotar la referencia o referencias que tienen problemas     Referencia    Control bloqueadoError    Guía de recolección    Número de guia con el que se va a recoger el pedido.     Guía de recolección    Control bloqueadoError    Pedido de reserva    Pedido con el cual el callcenter reserva las prendas del cliente.     Pedido de reserva    Control bloqueadoError    Cliente    Cliente Marca     Cliente           Select a value.  Control bloqueadoError               Auditoria    Grupo   Required Id. del propietario     Grupo  Chevignon         Select a value.  Control bloqueadoError    Agente asignado    Responsable de atender el caso     Agente asignado Agente CRM1         Select a value.  Control bloqueadoError      Escalar a    Referencia a listado de equipos     Escalar a           Select a value.  Control bloqueadoError                      Select a value.  Control bloqueadoError                     Select a value.  Control bloqueadoError                                    Select a value.  Control bloqueadoError                     Select a value.  Control bloqueadoError                     Select a value.  Control bloqueadoError                     Select a value.  Control bloqueadoError             Contraer esta ficha   Actividades         Descripción         Descripción   Control bloqueadoError                       ASISTENTEACTIVIDADESREGISTROS de KBNOTAS                        No hay acciones pendientes de completarse actualmente.           Todosdropdown      |Agregar Llamada de teléfonoMÁS COMANDOS Notificación Cita Tarea Llamada de teléfono Alerta de Customer Voice Canje de invitación Comentario Suscripción de alerta Ordenar por Todos                             Refresh the window                    Comentario       Público        Abrir esta actividad       Expandir esta actividad    Contraer esta actividad     ¡Hola!Recibimos tu solicitud identificada con el número CAS-255363-M3W6H9 y ha sido enviada al área encargada. Si desea agregar un comentario adicional puedes hacerlo respondiendo a este correo.   CHEVIGNON - Pregunta o solicitud. ¡Feliz Día! Chevignon       Completado por     SRV Admin Dynamics      Hoy       Comentario       Interno Portal        Abrir esta actividad       Expandir esta actividad    Contraer esta actividad    body>Marca: CHEVIGNON Nombre Completo: Carlos Javier Zapata MÃ¡rquez Email: cz5206159@gmail.com Solicitud: Pregunta o solicitud Celular: 3053129434 Tipo de Documento: Cedula Número de Documento: 94406346 Número de Pedido:  Descripción: Buenos días solicito muy amablemente el paz y salvo de mi crédito muchas gracias      Completado por     SRV Admin Dynamics      Hoy       Notificación       <Chevignon>    Abrir esta actividad       Expandir esta actividad    Contraer esta actividad   [Chevignon] Solicitud recibida CHEVIGNON - Pregunta o solicitud. CRM:0019803                   Hoy      Notificación       <Chevignon>    Abrir esta actividad       Expandir esta actividad    Contraer esta actividad   [Chevignon] Solicitud recibida CHEVIGNON - Pregunta o solicitud. CRM:0019800                   Hoy      Notificación       <Chevignon>    Abrir esta actividad       Expandir esta actividad    Contraer esta actividad   [Chevignon] Solicitud recibida CHEVIGNON - Pregunta o solicitud. CRM:0019802                   Hoy      Notificación       <Chevignon>    Abrir esta actividad       Expandir esta actividad    Contraer esta actividad   [Chevignon] Solicitud recibida CHEVIGNON - Pregunta o solicitud. CRM:0019801                   Hoy      Notificación       <Chevignon>    Abrir esta actividad       Expandir esta actividad    Contraer esta actividad   [Chevignon] Solicitud recibida CHEVIGNON - Pregunta o solicitud. CRM:0019799                   Hoy      Notificación       <Chevignon>    Abrir esta actividad       Expandir esta actividad    Contraer esta actividad   [Chevignon] Solicitud recibida CHEVIGNON - Pregunta o solicitud. CRM:0019798                   Hoy      Notificación          Abrir esta actividad       Expandir esta actividad    Contraer esta actividad   CHEVIGNON - Pregunta o solicitud.                   Hoy      Notificación          Abrir esta actividad       Expandir esta actividad    Contraer esta actividad   CHEVIGNON - Pregunta o solicitud.                   Hoy     Ver actividades más antiguas   Más Cargando...            search                 dropdown       Buscando...                                            Contraer esta ficha   Detalles del SLA         SLA APLICABLE         Control bloqueadoError                   Primera respuesta en   En curso 6d 6h 47m 33s         Control bloqueadoError      Control bloqueadoError         Control bloqueadoError                     Resolver en   En curso 13d 6h 47m 34s         Control bloqueadoError      Control bloqueadoError    Último período de retención    Contiene la marca de fecha y hora del último período de retención. Escriba una fecha con el formato d/MM/yyyy     Último período de retención          Seleccionar fecha        Seleccionar una hora                                                                                                       Control bloqueadoError    Período de retención (minutos)    Muestra la duración en minutos durante la que se retuvo el caso.     Período de retención (minutos) Muestra la duración en minutos durante la que se retuvo el caso.      Control bloqueado Error    CUMPLIMIENTO    CUMPLIMIENTO DEL SLA     CUMPLIMIENTO A TIEMPO   Control bloqueadoError    SEMAFORO    SEMAFORO PARA SABER EL ESTADO REAL DEL CASO     SEMAFORO VENCE MAS DE 5 DIAS   Control bloqueadoError                                                       NombreLos datos se almacenan en orden ascendente en esta columna        EstadoLos datos se almacenan en orden ascendente en esta columna        Hora del errorLos datos se almacenan en orden ascendente en esta columna        Hora de advertenciaLos datos se almacenan en orden ascendente en esta columna        Correcto elLos datos se almacenan en orden ascendente en esta columna                        Nombre  Estado  Hora del error  Hora de advertencia  Correcto el   Checkbox  Primera respuesta por KPI   En curso   25/11/2020 4:00 p.m.   24/11/2020 4:00 p.m.        Checkbox  Resolver por KPI   En curso   2/12/2020 4:00 p.m.   30/11/2020 4:00 p.m.                                                                                   Contraer esta ficha   Devolución de dinero         Geneal    Tipo de devolución    Metodo por el cual se le va realizar la devolución al cliente.     Tipo de devolución    Control bloqueadoError    Asesor que ingresa        Asesor que ingresa    Control bloqueadoError    Fecha de devolución o saldo    Fecha en la que se crea el sado el saldo a favor o se toman los datos para la devolución del dinero. Escriba una fecha con el formato d/MM/yyyy     Fecha de devolución o saldo          Seleccionar fecha        Seleccionar una hora                                                                                                       Control bloqueadoError    Motivos de devolución    Motivos por los cuales vamos a realizar la devolución al cliente     Motivos de devolución    Control bloqueadoError    Nombre cliente devolución    Nombre del cliente al cual se le va devolver el dinero     Nombre cliente devolución    Control bloqueadoError    Cédula cliente devolución    Cédula del cliente al cual se le va devolver el dinero     Cédula cliente devolución    Control bloqueadoError    Pedido a devolver    Pedido al cual le realizaremos al devolución del dinero     Pedido a devolver    Control bloqueadoError    Fecha del pedido    Fecha en la que el cliente realizo el pedido Escriba una fecha con el formato d/MM/yyyy     Fecha del pedido         Seleccionar fecha  Control bloqueadoError                  Banco a devolver    Banco de la devolución     Banco a devolver    Control bloqueadoError   Número de saldo    Es el codigo de saldo a favor que se le va entregar al cliente     Número de saldo    Control bloqueadoError    Tipo de cuenta    Tipo de cuenta al cual se va devolver el dinero     Tipo de cuenta    Control bloqueadoError   Vigencia de saldo a favor    Fecha de vencimiento del saldo a favor Escriba una fecha con el formato d/MM/yyyy     Vigencia de saldo a favor         Seleccionar fecha  Control bloqueadoError    Número de cuenta    Número de cuenta a la cual se le va devolver al cliente.     Número de cuenta    Control bloqueadoError                    Valor devolución o saldo    Valor a devolver al cliente     Valor devolución o saldo    Control bloqueadoError    ¿Tiene flete?    La devolución del dinero contiene flete.     ¿Tiene flete?    Control bloqueadoError    Valor del flete    Valor correspondiente al flete     Valor del flete    Control bloqueadoError    Observación    Observación sobre la devolución o saldo a favor.     Observación    Control bloqueadoError            Expandir esta ficha                           Control bloqueadoError              Control bloqueadoError                    Select a value.  Control bloqueadoError                 Control bloqueadoError                    Select a value.  Control bloqueadoError     Required               Select a value.  Control bloqueadoError                      Select a value.  Control bloqueadoError                    Seleccionar fecha        Seleccionar una hora                                                                                                       Control bloqueadoError                    Select a value.  Control bloqueadoError                      Select a value.  Control bloqueadoError     Recommended        Control bloqueadoError                      Seleccionar fecha        Seleccionar una hora                                                                                                        Control bloqueado Error                    Seleccionar fecha        Seleccionar una hora                                                                                                       Control bloqueadoError                               Select a value.  Control bloqueadoError                   Error               Agregar registro de Conexión.  Consultar los registros asociados a esta vista                        Los datos se almacenan en orden ascendente en esta columna                   0 Conexiones para Caso. Seleccione Agregar (+).                                                                                                                        Select a value.  Control bloqueadoError                  Iniciar búsquedaBorrar búsquedaHerramientas de búsqueda          Hacer clic aquí para obtener una vista normal                          Los datos se almacenan en orden ascendente en esta columna                           Checkbox        Checkbox        Checkbox        Checkbox        Checkbox        Checkbox        Checkbox        Checkbox                                                                  Cargar página siguiente              Seleccione un gráfico        Actualizar el gráfico       Cargando...                                Hacer clic aquí para obtener una vista normal                                                                                  Control bloqueadoError                          Control bloqueadoError              Recomendado         Control bloqueadoError              Control bloqueadoError     Recomendado         Control bloqueadoError                           Select a value.  Control bloqueadoError               Control bloqueadoError               Control bloqueadoError               Control bloqueadoError                              Hacer clic aquí para obtener una vista normal                            Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna                              Checkbox           Checkbox           Checkbox           Checkbox           Checkbox                                                                     Cargar página siguiente              Seleccione un gráfico        Actualizar el gráfico       Cargando...                                Hacer clic aquí para obtener una vista normal                                                                               Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna                            No se encontró ningún registro de Derecho.                                                                                                                                          Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna                                    Checkbox                  Checkbox                  Checkbox                  Checkbox                  Checkbox                                                                                                                       Error               Agregar registro de Caso.                               Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna        Los datos se almacenan en orden ascendente en esta columna                          0 Casos para Caso. Seleccione Agregar (+).                                                                                     Expandir esta ficha                                   Select a value.  Control bloqueadoError                                                 Select a value.  Control bloqueadoError                    Select a value.   Control bloqueado Error              Control bloqueadoError              Control bloqueadoError           Control bloqueadoError      Control bloqueadoError                                       Estado Obligatorio Activo   Control bloqueadoError                               Control bloqueadoError                                 Expandido, haga clic aquí para contraer                         Expandido, haga clic aquí para contraer                                                                                                                                                                                                                     
'''

    motivo,negocio,tipologia3,tipologia4,CPC_Escala,email = compare_text(message)
    print(motivo,negocio,tipologia3,tipologia4,CPC_Escala,email)


    
