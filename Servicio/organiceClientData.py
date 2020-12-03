'''
Resive el texto de la pagina web de Vtex que contiene entre otros datos, los datos del cliente,
y Retorna los datos del cliente organizados:
1.Nombre del Cliente
2.Celular
3.Correo Electronico[este campo depende]
4.Tipo de documento----> simpre que encuetre la palabra "CEDULACOL" sera "cedula" de lo contrario no se llena
5.Numero de Identificacion
6.Ciudad -------------> se utilizara un diccionario con todas las ciudades de Colombia para encontrar esta ciudad.
7.Numero de Pedido----------------> Numero entre parentesis en la pagina en la etiqueta Ventas
8.Guia de Despacho----------------> etiquetado como numero de Rastreo en la pagina
9.Referencia
'''
import re
import openpyxl

def read_dict_city():
    dictCity = openpyxl.load_workbook('ciudadesCol.xlsx')
    hoja1 = dictCity.get_sheet_by_name("Hoja1")

    listCity=[]
    for indexCell in range(1,hoja1.max_row):
        listCity.append(hoja1.cell(row=indexCell,column = 1).value)

    return listCity

def replace_accent(text):
    # Quita las tildes del texto que ingresa
    text = text.replace("á","a")
    text = text.replace("é","e")
    text = text.replace("í","i")
    text = text.replace("ó","o")
    text = text.replace("ú","u")
    return text

def extract_name(lineText):
    # Resive la linea donde se encuntra el nombre del Cliente
    # Retorna solamente el nombre del Cliente
    stringSearch="Direccion de Entrega"
    indexIniAddress=lineText.find(stringSearch)
    indexIniName = indexIniAddress + len(stringSearch)+1
    name = lineText[indexIniName:]
    return name

def extract_cellphone(lineText):
    # Resive la linea donde se encuntra el numero de celular
    # Retorna solamente el numero de celular
    stringSearch="CEDULACOL"
    numCelular=lineText[0:lineText.find(stringSearch)]
    numCelular=re.sub("\D","",numCelular)
    return numCelular


def extract_idcard(lineText):
    # Resive la linea donde se encuntra la cedula
    # Retorna solamente el numero de cedula
    stringSearch="CEDULACOL"
    idcard=lineText[lineText.find(stringSearch)+len(stringSearch)+1:]
    idcard=re.sub("\D","",idcard)
    return idcard   

def extract_city(lineCity):
    # Resive la line en donde se encuntra la direccion y la ciudad
    # Filtra la ciudad y la retorna
    lineCity=replace_accent(lineCity)
    #print(lineCity)
    listCities=read_dict_city()
    for city in listCities:
        if city.lower() in lineCity.lower():
            #print(city)
            break

    return city

def extract_num_order(line):
    # Resive la linea donde se encuentra el numero de pedido entre parentecis
    # Filtra y retorna el numero de orden 
    indexNumOrderIni=line.find("(")
    indexNumOrdedEnd=line.find(")")
    numOrder=line[indexNumOrderIni+1:indexNumOrdedEnd]
    return numOrder

def extract_office_guide(line):
    # Resive la linea donde se encuentra el numero de la Guia de la transportadora 
    # Filtra y retorna el numero de guia
    officeGuide=re.sub("\D","",line)
    return officeGuide


def organice_client_data(textPage):
    textPage = replace_accent(textPage)
    textPageList=textPage.split("  ")
    name,cellPhone,idCard,city,numOrder,officeGuide=" "," "," "," "," "," "
    for indexLine,line in enumerate(textPageList):
        print(indexLine,line)
        if "Direccion de Entrega" in line:
            city=extract_city(textPageList[indexLine+1])
            name=extract_name(line)
        elif "CEDULACOL" in line:
            cellPhone=extract_cellphone(line)
            idCard=extract_idcard(line)
        elif "Ventas" in line:
            numOrder=extract_num_order(textPageList[indexLine+2])
        elif "Numero de rastreo" in line:
            officeGuide=extract_office_guide(line)
            print("-------------------------------Guia Transportadora:",officeGuide)
            
    
    return (name,cellPhone,"Cédula",idCard,city,numOrder,officeGuide)
        

if __name__ == "__main__":
    textPage=" Ventas    1079901467559-01 (731490)  REALIZADA EN 28 NOV 2020 17:24 Vendido por lukshop Imprimir  DATOS DEL CLIENTE Leidy... Fuentes angelik87garcia@gmail.com   312 395 9464 CEDULACOL: 1121840529    Reenviar el último email VALOR TOTAL 27.930 COP Total de los items39.900 COP  Total de los Descuentos-11.970 COP  PROMOCIONES E MARKETING Promociones blacksale_30_20201126 blacksale_envio_20201128  ESTADO DEL PEDIDO Facturado ESTADO DE LA TRANSACIÓN Su perfil no tiene acceso a esta información.  FLUJOMOSTRAR INTERACCIONES Pedido aceptadoFinal del períodoItens a facturarTodos itens facturadosAceptadoNegadoEsperando confirmación del sellerPago pendiente28/11/2017:25Pago aprobado28/11/2017:26Tiempo para cancelación28/11/2017:26Listo para preparar28/11/2017:26A preparar la entrega28/11/2017:26Cancelación solicitadaFacturado30/11/2010:02Aprobar pago28/11/2017:25Autorizar envío28/11/2017:25Empezar la preparación28/11/2017:26Verifica factura30/11/2010:02Pago negadoSolicitar cancelamentoSolicitar cancelamentoEmpezando cancelaciónCancelado FACTURA  Factura No. VON5538511 Fecha de Emisíon: 29 Nov 2020 19:00  Valor: 27.930 COP  Paquete 1 (1 producto, 1 unidad) Item  Camisa estampado hojas Azul S 25601 (ref 36_210B003_7705261852681) Almacén: 1_1 1 Item 27.930 COP Total de los items  27.930 COP Forma de entrega Envío  Estimado - 11 dic 2020 00:00 COORDINADORA  Número de rastreo: 48871244717  Editar datos de rastreo  Dirección de Entrega Leidy Angelica García Fuentes  Calle 13#12b Estero Villavicencio / Meta 50001 – COL PEDIDO APROBADO Pedido a ser cobrado  28 nov 2020 17:25 PAGO BANCO DAVIVIENDA 27.930 COP 1 cuotas de 27.930 COP  Autorización de Gateway 28 nov 2020 17:25  ID del pago 13F3368DCCD14B79B863E78A0D8F81A5  TID del pago 814199200  Transación ID: BD9C1EB06E3844E091CC41C6AD99D35F Merchant: LUKSHOP Ver detalles de la transacción  LÍNEA DE TIEMPO 28 nov 2020 17:24:28 Criado 17:24:28 Pedido aceptado 17:24:28 Procesando 17:24:29 En espera de la autorización para despachar 17:25:47 Criado (incompleto) 17:25:47 Procesando 17:25:49 Pago pendiente 17:25:52 RIFLE servicioalcliente@rifle.com.co Su pedido Camisa estampado hoj... fue exitoso!   Para angelik87garcia@gmail.com      Reenviar email  17:26:00 Tiempo para cancelación 17:26:00 Pago aprobado 17:26:03 RIFLE servicioalcliente@rifle.com.co Confirmación de pago del pedido 1079901467559-01   Para angelik87garcia@gmail.com      Reenviar email  17:26:04 Listo para preparar 17:26:33 Empezar la preparación 17:26:33 A preparar la entrega 30 nov 2020 10:02:20 Verifica factura 10:02:21 Facturado 10:02:21 RIFLE servicioalcliente@rifle.com.co El pedido 1079901467559-01 ya fue facturado   Para angelik87garcia@gmail.com      Reenviar email  10:02:27 RIFLE servicioalcliente@rifle.com.co Tu pedido 1079901467559-01 ya fue despachado   Para angelik87garcia@gmail.com      Reenviar email  Ahora Tu comentario  Los comentarios no se envían a los clientes, adelante.  Comentar pedido  TODAS295 Buscar por nombre, persona, comentario... Catálogo Comercial Financiero Infraestructura Backoffice & ERP Frontend Marketplace"
    organice_client_data(textPage)